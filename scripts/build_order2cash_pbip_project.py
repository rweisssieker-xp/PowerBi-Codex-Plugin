#!/usr/bin/env python3
"""Generate a PBIP/TMDL-ready Order2Cash project scaffold."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import re
import shutil
import uuid
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
TEST_PACK = ROOT / "outputs" / "powerbi-order2cash-test"
OUT_DIR = ROOT / "outputs" / "powerbi-order2cash-pbip" / "Order2Cash"
SOURCE_XLSX = ROOT / "outputs" / "powerbi-industrial-demo" / "powerbi_industrial_demo_data.xlsx"
REPORT_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/report/2.0.0/schema.json"
PAGE_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/page/2.0.0/schema.json"
VISUAL_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/visualContainer/2.0.0/schema.json"
PBIR_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/report/definitionProperties/2.0.0/schema.json"
PBISM_SCHEMA = "https://developer.microsoft.com/json-schemas/fabric/item/semanticModel/definitionProperties/1.0.0/schema.json"
MEASURE_TABLE = "Fact_SalesOrders"
SOURCE_MODE = "embedded"
EXCEL_SHEET_NAMES: set[str] | None = None
MEASURE_VISUAL_FALLBACKS = {
    "Order Count": ("Fact_SalesOrders", "SalesOrderID"),
    "Order Value": ("Fact_SalesOrders", "OrderValue"),
    "Open Backlog Value": ("Fact_SalesOrders", "OrderValue"),
    "ATP Issue Rate": ("Fact_SalesOrders", "ATPStatus"),
    "Late Or Short Delivery Rate": ("Fact_Deliveries", "OTIFStatus"),
    "Late Or Short Deliveries": ("Fact_Deliveries", "DeliveryID"),
    "Invoice Amount": ("Fact_Billing", "InvoiceAmount"),
    "Open AR Amount": ("Fact_AR", "OpenAmount"),
    "Billing Block Rate": ("Fact_Billing", "BillingBlockFlag"),
    "Quote Value": ("Fact_Quotes", "QuoteValue"),
    "Discount Leakage": ("Fact_Quotes", "DiscountPct"),
    "Churn Risk Customers": ("Dim_Customer", "CustomerID"),
    "O2C Exception Count": ("Fact_Order2CashExceptions", "ExceptionID"),
    "O2C Exception Exposure": ("Fact_Order2CashExceptions", "AmountExposure"),
}


TABLES = [
    "Dim_Calendar",
    "Dim_Customer",
    "Dim_Product",
    "Dim_Plant",
    "Dim_Warehouse",
    "Fact_Leads",
    "Fact_Opportunities",
    "Fact_Quotes",
    "Fact_SalesOrders",
    "Fact_Deliveries",
    "Fact_Billing",
    "Fact_AR",
    "Fact_ServiceTickets",
    "Fact_Shipments",
    "Fact_GL_Postings",
    "Fact_Order2CashExceptions",
]


COLUMNS = {
    "Dim_Customer": ["CustomerID", "CustomerName", "Segment", "Industry", "Country", "SalesRegion", "RiskClass", "ActiveFlag"],
    "Dim_Product": ["ProductID", "ProductName", "ProductFamily", "ProductLine", "LifecycleStatus", "StandardCost", "ListPrice"],
    "Dim_Calendar": ["Date", "Year", "Quarter", "MonthNo", "MonthName", "FiscalYear", "FiscalPeriod", "IsWorkingDay"],
    "Dim_Plant": ["PlantID", "CompanyCode", "PlantName", "Country", "IndustryFocus"],
    "Dim_Warehouse": ["WarehouseID", "PlantID", "WarehouseType", "WarehouseName"],
    "Fact_Leads": ["LeadID", "CreatedDate", "Source", "Industry", "Country", "LeadScore", "Status", "OwnerID"],
    "Fact_Opportunities": ["OpportunityID", "LeadID", "CustomerID", "CreatedDate", "Stage", "ProbabilityPct", "ExpectedValue", "ExpectedCloseDate", "SalesOwnerID"],
    "Fact_ServiceTickets": ["ServiceTicketID", "CustomerID", "ProductID", "InstalledBaseID", "OpenDate", "CloseDate", "Priority", "SLAStatus", "WarrantyFlag", "ServiceCost"],
    "Fact_Shipments": ["ShipmentID", "DeliveryID", "Carrier", "Mode", "PlannedPickup", "ActualPickup", "PlannedDelivery", "ActualDelivery", "FreightCost", "CO2eKg"],
    "Fact_GL_Postings": ["GLDocumentID", "CompanyCode", "PostingDate", "FiscalYear", "GLAccount", "CostCenter", "ProfitCenter", "Amount", "Currency", "SourceProcess"],
    "Fact_SalesOrders": ["SalesOrderID", "QuoteID", "CustomerID", "ProductID", "PlantID", "OrderDate", "RequestedDate", "ConfirmedDate", "OrderQty", "NetPrice", "OrderValue", "Status", "ATPStatus"],
    "Fact_Deliveries": ["DeliveryID", "SalesOrderID", "CustomerID", "ProductID", "WarehouseID", "PickDate", "ShipDate", "DeliveryDate", "DeliveredQty", "OTIFStatus", "Carrier"],
    "Fact_Billing": ["InvoiceID", "DeliveryID", "SalesOrderID", "CustomerID", "InvoiceDate", "DueDate", "InvoiceAmount", "TaxAmount", "BillingBlockFlag"],
    "Fact_AR": ["InvoiceID", "CustomerID", "DueDate", "OpenAmount", "PaymentDate", "DisputeFlag", "DunningLevel"],
    "Fact_Quotes": ["QuoteID", "OpportunityID", "CustomerID", "QuoteDate", "QuoteStatus", "QuoteValue", "DiscountPct", "ApprovalCycleDays", "LegalLoopCount"],
    "Fact_Order2CashExceptions": ["ExceptionID", "ProblemType", "Severity", "CustomerID", "ProductID", "RelatedDocument", "EventDate", "AmountExposure", "Reason", "RecommendedAction", "OwnerRole"],
}


RELATIONSHIPS = [
    ("Fact_SalesOrders", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_SalesOrders", "ProductID", "Dim_Product", "ProductID"),
    ("Fact_SalesOrders", "PlantID", "Dim_Plant", "PlantID"),
    ("Fact_Deliveries", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_Deliveries", "ProductID", "Dim_Product", "ProductID"),
    ("Fact_Deliveries", "WarehouseID", "Dim_Warehouse", "WarehouseID"),
    ("Fact_Billing", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_AR", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_Quotes", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_ServiceTickets", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_ServiceTickets", "ProductID", "Dim_Product", "ProductID"),
    ("Fact_Order2CashExceptions", "CustomerID", "Dim_Customer", "CustomerID"),
    ("Fact_Order2CashExceptions", "ProductID", "Dim_Product", "ProductID"),
]


PAGE_SPECS = [
    {
        "name": "p01_o2c_overview",
        "displayName": "O2C Executive Overview",
        "visuals": [
            ("v01_order_value", "card", "Order Value", "_Measures", "Order Value", 24, 24, 210, 110),
            ("v02_backlog", "card", "Open Backlog Value", "_Measures", "Open Backlog Value", 250, 24, 210, 110),
            ("v03_exception_count", "card", "O2C Exception Count", "_Measures", "O2C Exception Count", 476, 24, 210, 110),
            ("v04_exception_exposure", "card", "O2C Exception Exposure", "_Measures", "O2C Exception Exposure", 702, 24, 240, 110),
            ("v05_problem_mix", "clusteredBarChart", "Exceptions by problem type", "Fact_Order2CashExceptions", "ProblemType", 24, 170, 440, 300),
            ("v06_action_backlog", "tableEx", "Top O2C exception backlog", "Fact_Order2CashExceptions", "AmountExposure", 490, 170, 650, 300),
        ],
    },
    {
        "name": "p02_customer_churn",
        "displayName": "Customer Churn and Retention Risk",
        "visuals": [
            ("v01_churn_customers", "card", "Churn Risk Customers", "_Measures", "Churn Risk Customers", 24, 24, 220, 110),
            ("v02_customer_score", "tableEx", "Customer risk score detail", "Fact_Order2CashExceptions", "CustomerID", 24, 160, 620, 330),
            ("v03_customer_revenue", "clusteredBarChart", "High exposure churn customers", "Fact_Order2CashExceptions", "AmountExposure", 670, 160, 470, 330),
        ],
    },
    {
        "name": "p03_order_backlog",
        "displayName": "Order Intake and Backlog",
        "visuals": [
            ("v01_order_count", "card", "Order Count", "_Measures", "Order Count", 24, 24, 220, 110),
            ("v02_atp_rate", "card", "ATP Issue Rate", "_Measures", "ATP Issue Rate", 260, 24, 220, 110),
            ("v03_order_status", "clusteredColumnChart", "Order status and ATP", "Fact_SalesOrders", "Status", 24, 160, 540, 320),
            ("v04_backlog_table", "tableEx", "Backlog exceptions", "Fact_Order2CashExceptions", "RelatedDocument", 590, 160, 550, 320),
        ],
    },
    {
        "name": "p04_delivery_otif",
        "displayName": "Delivery and OTIF",
        "visuals": [
            ("v01_delivery_rate", "card", "Late Or Short Delivery Rate", "_Measures", "Late Or Short Delivery Rate", 24, 24, 260, 110),
            ("v02_late_count", "card", "Late Or Short Deliveries", "_Measures", "Late Or Short Deliveries", 300, 24, 230, 110),
            ("v03_carrier_otif", "clusteredBarChart", "Carrier and delivery exceptions", "Fact_Deliveries", "Carrier", 24, 160, 540, 320),
            ("v04_otif_backlog", "tableEx", "Late/short action backlog", "Fact_Order2CashExceptions", "ProblemType", 590, 160, 550, 320),
        ],
    },
    {
        "name": "p05_billing_cash",
        "displayName": "Billing and Cash",
        "visuals": [
            ("v01_invoice_amount", "card", "Invoice Amount", "_Measures", "Invoice Amount", 24, 24, 220, 110),
            ("v02_open_ar", "card", "Open AR Amount", "_Measures", "Open AR Amount", 260, 24, 220, 110),
            ("v03_billing_block", "card", "Billing Block Rate", "_Measures", "Billing Block Rate", 496, 24, 220, 110),
            ("v04_cash_exceptions", "tableEx", "Cash risk and disputes", "Fact_Order2CashExceptions", "AmountExposure", 24, 160, 1110, 340),
        ],
    },
    {
        "name": "p06_quote_margin",
        "displayName": "Quote and Margin Leakage",
        "visuals": [
            ("v01_quote_value", "card", "Quote Value", "_Measures", "Quote Value", 24, 24, 220, 110),
            ("v02_discount_leakage", "card", "Discount Leakage", "_Measures", "Discount Leakage", 260, 24, 220, 110),
            ("v03_approval_aging", "clusteredColumnChart", "Approval cycle and legal loops", "Fact_Quotes", "ApprovalCycleDays", 24, 160, 540, 320),
            ("v04_quote_table", "tableEx", "Quote leakage detail", "Fact_Quotes", "DiscountPct", 590, 160, 550, 320),
        ],
    },
]


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def tmdl_string(value: Path | str) -> str:
    return str(value).replace("\\", "/").replace('"', '\\"')


def tmdl_identifier(value: str) -> str:
    if re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", value):
        return value
    return "'" + value.replace("'", "''") + "'"


def literal(value: str) -> dict[str, object]:
    return {"expr": {"Literal": {"Value": value}}}


def text_literal(value: str) -> dict[str, object]:
    return literal("'" + value.replace("'", "''") + "'")


def color_literal(value: str) -> dict[str, object]:
    return {"solid": {"color": literal("'" + value + "'")}}


def stable_guid(value: str) -> str:
    return str(uuid.uuid5(uuid.NAMESPACE_URL, value))


def pbir_name(prefix: str, text: str, max_slug: int = 18) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "", text.title())
    return f"{prefix}{slug[:max_slug]}"


def data_type(column: str) -> str:
    if column.endswith("Flag"):
        return "boolean"
    if column.endswith("Date") or column == "Date" or column in {"PlannedPickup", "ActualPickup", "PlannedDelivery", "ActualDelivery"}:
        return "dateTime"
    if column in {
        "OrderQty",
        "DeliveredQty",
        "DunningLevel",
        "Year",
        "MonthNo",
        "FiscalYear",
        "FiscalPeriod",
        "ApprovalCycleDays",
        "LegalLoopCount",
        "LeadScore",
    }:
        return "int64"
    if (
        any(token in column for token in ["Amount", "Value", "Price", "Pct", "Exposure"])
        or column in {"StandardCost", "ServiceCost", "FreightCost", "CO2eKg"}
    ):
        return "double"
    return "string"


def datatable_type(column: str, typed: bool = True) -> str:
    if not typed:
        return "STRING"
    mapping = {
        "boolean": "BOOLEAN",
        "dateTime": "DATETIME",
        "int64": "INTEGER",
        "double": "DOUBLE",
        "string": "STRING",
    }
    return mapping[data_type(column)]


def dax_value(value: object, column_type: str) -> str:
    if value is None or value == "":
        if column_type == "BOOLEAN":
            return "FALSE"
        if column_type == "DATETIME":
            return 'dt"1900-01-01"'
        if column_type in {"INTEGER", "DOUBLE"}:
            return "0"
        return '""'
    if column_type == "BOOLEAN":
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        return "TRUE" if str(value).strip().lower() in {"true", "1", "yes", "y"} else "FALSE"
    if column_type == "DATETIME":
        if isinstance(value, dt.datetime):
            return f'dt"{value.year:04d}-{value.month:02d}-{value.day:02d}"'
        if isinstance(value, dt.date):
            return f'dt"{value.year:04d}-{value.month:02d}-{value.day:02d}"'
        text = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y"):
            try:
                parsed = dt.datetime.strptime(text[:10], fmt)
                return f'dt"{parsed.year:04d}-{parsed.month:02d}-{parsed.day:02d}"'
            except ValueError:
                pass
        return '"' + text.replace('"', '""') + '"'
    if column_type in {"INTEGER", "DOUBLE"}:
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, (int, float)):
            return str(int(value)) if column_type == "INTEGER" else repr(float(value))
        text = str(value).strip().replace(",", ".")
        if not text:
            return "BLANK()"
        return str(int(float(text))) if column_type == "INTEGER" else repr(float(text))
    return '"' + str(value).replace('"', '""') + '"'


def read_excel_sheet(table: str) -> tuple[list[str], list[dict[str, object]]]:
    workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    try:
        sheet = workbook[table]
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value) for value in next(rows)]
        records = []
        for row in rows:
            records.append({header: row[idx] if idx < len(row) else None for idx, header in enumerate(headers)})
        return headers, records
    finally:
        workbook.close()


def excel_sheet_names() -> set[str]:
    global EXCEL_SHEET_NAMES
    if EXCEL_SHEET_NAMES is None:
        workbook = load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
        try:
            EXCEL_SHEET_NAMES = set(workbook.sheetnames)
        finally:
            workbook.close()
    return EXCEL_SHEET_NAMES


def read_csv_table(table: str) -> tuple[list[str], list[dict[str, object]]]:
    path = TEST_PACK / "data" / f"{table}.csv"
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def table_records(table: str) -> tuple[list[str], list[dict[str, object]]]:
    csv_path = TEST_PACK / "data" / f"{table}.csv"
    if csv_path.exists():
        return read_csv_table(table)
    return read_excel_sheet(table)


def datatable_source(table: str, columns: list[str], records: list[dict[str, object]], typed: bool) -> list[str]:
    lines = ["DATATABLE("]
    for idx, column in enumerate(columns):
        suffix = "," if idx < len(columns) - 1 or records else ""
        lines.append(f'\t"{column}", {datatable_type(column, typed)}{suffix}')
    lines.append("\t{")
    for row_idx, record in enumerate(records):
        values = [dax_value(record.get(column), datatable_type(column, typed)) for column in columns]
        suffix = "," if row_idx < len(records) - 1 else ""
        lines.append(f"\t\t{{ {', '.join(values)} }}{suffix}")
    lines.append("\t}")
    lines.append(")")
    return lines


def native_excel_source(table: str) -> list[str]:
    type_pairs = ", ".join(f'{{"{column}", {powerquery_type(column)}}}' for column in COLUMNS[table])
    return [
        "let",
        f'    Source = Excel.Workbook(File.Contents("{tmdl_string(SOURCE_XLSX)}"), null, true),',
        f'    Raw = Source{{[Item="{table}", Kind="Sheet"]}}[Data],',
        "    PromotedHeaders = Table.PromoteHeaders(Raw, [PromoteAllScalars = true]),",
        f"    ChangedType = Table.TransformColumnTypes(PromotedHeaders, {{{type_pairs}}}, \"en-US\")",
        "in",
        "    ChangedType",
    ]


def native_csv_source(table: str) -> list[str]:
    csv_path = TEST_PACK / "data" / f"{table}.csv"
    type_pairs = ", ".join(f'{{"{column}", {powerquery_type(column)}}}' for column in COLUMNS[table])
    return [
        "let",
        f'    Source = Csv.Document(File.Contents("{tmdl_string(csv_path)}"), [Delimiter = ",", Encoding = 65001, QuoteStyle = QuoteStyle.Csv]),',
        "    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars = true]),",
        f"    ChangedType = Table.TransformColumnTypes(PromotedHeaders, {{{type_pairs}}}, \"en-US\")",
        "in",
        "    ChangedType",
    ]


def native_source(table: str) -> list[str]:
    if table in excel_sheet_names():
        return native_excel_source(table)
    if (TEST_PACK / "data" / f"{table}.csv").exists():
        return native_csv_source(table)
    raise RuntimeError(f"No native source found for {table}: missing Excel sheet and CSV file")


def powerquery_type(column: str) -> str:
    mapping = {
        "boolean": "type logical",
        "dateTime": "type date",
        "int64": "Int64.Type",
        "double": "type number",
        "string": "type text",
    }
    return mapping[data_type(column)]


def table_tmdl(table: str) -> str:
    headers, records = table_records(table)
    cols = COLUMNS.get(table, [])
    typed = bool(cols)
    if not cols:
        cols = headers
    lines = [f"table {tmdl_identifier(table)}", f"\tlineageTag: {stable_guid(f'order2cash-table-{table}')}", ""]
    for column in cols:
        lines.extend(
            [
                f"\tcolumn {tmdl_identifier(column)}",
                f"\t\tdataType: {data_type(column) if typed else 'string'}",
                # Power BI Desktop writes DATATABLE source columns as DAX-style references.
                # The generic TOM serializer accepts bare names, but PBIP loading does not.
                f"\t\tsourceColumn: [{column}]",
                "",
            ]
        )
    partition_kind = "calculated" if SOURCE_MODE == "embedded" else "m"
    source_lines = datatable_source(table, cols, records, typed) if SOURCE_MODE == "embedded" else native_source(table)
    lines.extend(
        [
            f"\tpartition {tmdl_identifier(table)} = {partition_kind}",
            "\t\tmode: import",
            "\t\tsource =",
        ]
    )
    for line in source_lines:
        lines.append(f"\t\t\t{line}")
    lines.append("")
    if table == MEASURE_TABLE:
        lines.extend(measure_lines())
    return "\n".join(lines)


def measure_lines() -> list[str]:
    raw = (TEST_PACK / "dax" / "order2cash_measures.dax").read_text(encoding="utf-8")
    blocks: list[tuple[str, str]] = []
    current_name: str | None = None
    current_expr: list[str] = []
    for line in raw.splitlines():
        if not line.strip() or line.strip().startswith("--"):
            continue
        stripped = line.strip()
        is_measure_header = (
            line.endswith("=")
            and not line.startswith((" ", "\t"))
            and not stripped.upper().startswith(("VAR ", "RETURN"))
        )
        if is_measure_header:
            if current_name:
                blocks.append((current_name, "\n".join(current_expr).strip()))
            current_name = line[:-1].strip()
            current_expr = []
        else:
            current_expr.append(line)
    if current_name:
        blocks.append((current_name, "\n".join(current_expr).strip()))

    lines: list[str] = []
    for name, expr in blocks:
        lines.append(f"\tmeasure '{name}' =")
        for line in expr.splitlines():
            lines.append(f"\t\t\t{line}")
        lines.append("\t\tformatString: #,0.00")
        lines.append("")
    return lines


def relationships_tmdl() -> str:
    lines = []
    for idx, (from_table, from_col, to_table, to_col) in enumerate(RELATIONSHIPS, start=1):
        lines.extend(
            [
                f"relationship {uuid.uuid5(uuid.NAMESPACE_URL, f'order2cash-rel-{idx}-{from_table}-{from_col}-{to_table}-{to_col}')}",
                f"\tfromColumn: {from_table}.{from_col}",
                f"\ttoColumn: {to_table}.{to_col}",
                "",
            ]
        )
    return "\n".join(lines)


def database_tmdl() -> str:
    return f"""database {uuid.uuid5(uuid.NAMESPACE_URL, "order2cash-demo-semantic-model")}
\tcompatibilityLevel: 1600
\tcompatibilityMode: powerBI
\tlanguage: 1033
"""


def model_tmdl() -> str:
    table_refs = "\n".join(f"ref table {table}" for table in TABLES)
    return f"""model Order2Cash
\tculture: en-US
\tdefaultPowerBIDataSourceVersion: powerBI_V3
\tsourceQueryCulture: en-US
\tdataAccessOptions
\t\tlegacyRedirects
\t\treturnErrorValuesAsNull

annotation SourceWorkbook = "{tmdl_string(SOURCE_XLSX)}"
annotation BusinessProcess = "Order2Cash"
annotation Documentation = "Generated PBIP/TMDL scaffold from industrial demo data"

{table_refs}

ref cultureInfo en-US
"""


def visual_json(name: str, visual_type: str, title: str, table: str, field: str, x: int, y: int, width: int, height: int) -> str:
    """Create a PBIR visual descriptor with both metadata and generation annotations."""
    query_table, query_field = MEASURE_VISUAL_FALLBACKS.get(field, (table, field)) if table == "_Measures" else (table, field)
    field_projection = {
        "field": {
            "Column": {
                "Expression": {"SourceRef": {"Entity": query_table}},
                "Property": query_field,
            }
        },
        "queryRef": f"{query_table}.{query_field}",
        "nativeQueryRef": query_field,
        "displayName": title,
    }
    query_state = {"Values": {"projections": [field_projection]}}
    payload = {
        "$schema": VISUAL_SCHEMA,
        "name": name,
        "position": {
            "x": x,
            "y": y,
            "z": 0,
            "height": height,
            "width": width,
            "tabOrder": 0,
        },
        "visual": {
            "visualType": "tableEx",
            "query": {"queryState": query_state},
            "drillFilterOtherVisuals": True,
            "visualContainerObjects": {
                "title": [
                    {
                        "properties": {
                            "show": literal("true"),
                            "text": text_literal(title),
                            "fontSize": literal("12D"),
                            "bold": literal("true"),
                            "fontColor": color_literal("#1D2939"),
                            "titleWrap": literal("true"),
                        }
                    }
                ],
                "background": [{"properties": {"show": literal("true"), "color": color_literal("#FFFFFF"), "transparency": literal("0D")}}],
                "border": [{"properties": {"show": literal("true"), "color": color_literal("#D0D5DD"), "radius": literal("4D")}}],
                "visualHeader": [{"properties": {"show": literal("false")}}],
            },
        },
        "annotations": [
            {"name": "generatedBy", "value": "build_order2cash_pbip_project.py"},
            {"name": "businessPurpose", "value": title},
            {"name": "sourceTable", "value": table},
            {"name": "sourceFieldOrMeasure", "value": field},
        ],
    }
    return json.dumps(payload, indent=2)


def page_json(page: dict[str, object], ordinal: int, page_name: str) -> str:
    payload = {
        "$schema": PAGE_SCHEMA,
        "name": page_name,
        "displayName": page["displayName"],
        "displayOption": "FitToPage",
        "height": 720,
        "width": 1280,
        "annotations": [
            {"name": "ordinal", "value": str(ordinal)},
            {"name": "generatedBy", "value": "build_order2cash_pbip_project.py"},
            {"name": "businessProcess", "value": "Order2Cash"},
        ],
    }
    return json.dumps(payload, indent=2)


def pages_json() -> str:
    page_order = [
        pbir_name(f"ReportSection{idx:02d}", str(page["displayName"]))
        for idx, page in enumerate(PAGE_SPECS, start=1)
    ]
    payload = {
        "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/pagesMetadata/1.0.0/schema.json",
        "pageOrder": page_order,
        "activePageName": page_order[0],
    }
    return json.dumps(payload, indent=2)


def report_json() -> str:
    payload = {
        "$schema": REPORT_SCHEMA,
        "themeCollection": {
            "baseTheme": {
                "name": "CY24SU06",
                "type": "SharedResources",
                "reportVersionAtImport": "5.59",
            }
        },
        "annotations": [
            {"name": "defaultPage", "value": pbir_name("ReportSection01", str(PAGE_SPECS[0]["displayName"]))},
            {"name": "generatedBy", "value": "build_order2cash_pbip_project.py"},
        ],
    }
    return json.dumps(payload, indent=2)


def write_report_pages(report_definition: Path) -> None:
    write(
        report_definition / "version.json",
        json.dumps(
            {
                "$schema": "https://developer.microsoft.com/json-schemas/fabric/item/report/definition/versionMetadata/1.0.0/schema.json",
                "version": "2.0.0",
            },
            indent=2,
        ),
    )
    write(report_definition / "report.json", report_json())
    write(report_definition / "pages" / "pages.json", pages_json())
    for ordinal, page in enumerate(PAGE_SPECS):
        page_name = pbir_name(f"ReportSection{ordinal + 1:02d}", str(page["displayName"]))
        page_dir = report_definition / "pages" / page_name
        write(page_dir / "page.json", page_json(page, ordinal, page_name))
        for visual_idx, visual in enumerate(page["visuals"], start=1):
            visual_name, visual_type, title, table, field, x, y, width, height = visual
            visual_id = f"Visual{ordinal + 1:02d}{visual_idx:02d}"
            visual_dir = page_dir / "visuals" / visual_id
            write(visual_dir / "visual.json", visual_json(visual_id, visual_type, title, table, field, x, y, width, height))


def validate_tmdl_datatable_constants(project: Path) -> None:
    """Fail generation on DATATABLE row expressions Power BI rejects as non-constant."""
    definition = project / "Order2Cash.SemanticModel" / "definition"
    forbidden_patterns = {
        "DATE(": "use dt\"YYYY-MM-DD\" literals in DATATABLE rows",
        "TRUE()": "use TRUE literals in DATATABLE rows",
        "FALSE()": "use FALSE literals in DATATABLE rows",
        "BLANK()": "use typed fallback literals in DATATABLE rows",
        "NaN": "DATATABLE rows must contain finite numeric literals",
        "Infinity": "DATATABLE rows must contain finite numeric literals",
    }
    errors: list[str] = []
    for tmdl_file in definition.glob("**/*.tmdl"):
        text = tmdl_file.read_text(encoding="utf-8")
        if "DATATABLE(" not in text:
            continue
        for pattern, hint in forbidden_patterns.items():
            line_no = text.find(pattern)
            if line_no >= 0:
                line = text[:line_no].count("\n") + 1
                errors.append(f"{tmdl_file}:{line}: {pattern} is not safe here; {hint}")

    if errors:
        raise RuntimeError("Invalid DATATABLE constants:\n" + "\n".join(errors))


def validate_semantic_model_structure(project: Path) -> None:
    """Fail generation if the TMDL shape drifts away from a clean Power BI model."""
    table_dir = project / "Order2Cash.SemanticModel" / "definition" / "tables"
    errors: list[str] = []

    for table in TABLES:
        if table not in COLUMNS:
            errors.append(f"{table}: missing typed schema in COLUMNS")

    for table_file in table_dir.glob("*.tmdl"):
        text = table_file.read_text(encoding="utf-8")
        for source_column in re.finditer(r"^\s+sourceColumn:\s+(?P<value>.+)$", text, flags=re.MULTILINE):
            value = source_column.group("value").strip()
            if not re.fullmatch(r"\[[^\]]+\]", value):
                errors.append(f"{table_file}: sourceColumn must use Power BI PBIP DATATABLE form [ColumnName]")
        for match in re.finditer(r"^\tmeasure\s+'[^']+'\s*=\n(?P<body>.*?)(?=^\t(?:measure|column|partition)\s|\Z)", text, flags=re.MULTILINE | re.DOTALL):
            body = match.group("body")
            if re.search(r"^\t\t\tformatString:", body, flags=re.MULTILINE):
                errors.append(f"{table_file}: formatString is indented as DAX expression content")
            if not re.search(r"^\t\tformatString:", body, flags=re.MULTILINE):
                errors.append(f"{table_file}: measure is missing a parsed formatString property")

    model_columns: dict[str, set[str]] = {}
    for table_file in table_dir.glob("*.tmdl"):
        text = table_file.read_text(encoding="utf-8")
        table_match = re.search(r"^table\s+(.+)$", text, flags=re.MULTILINE)
        if not table_match:
            continue
        table_name = table_match.group(1).strip().strip("'")
        model_columns[table_name] = {
            match.group(1).strip().strip("'")
            for match in re.finditer(r"^\s+column\s+(.+)$", text, flags=re.MULTILINE)
        }

    for from_table, from_col, to_table, to_col in RELATIONSHIPS:
        if not from_table.startswith("Fact_") or not to_table.startswith("Dim_"):
            errors.append(
                "relationship topology must be star-schema Fact_* -> Dim_* only: "
                f"{from_table}.{from_col} -> {to_table}.{to_col}"
            )
        if from_col not in model_columns.get(from_table, set()):
            errors.append(f"relationship source missing: {from_table}.{from_col}")
        if to_col not in model_columns.get(to_table, set()):
            errors.append(f"relationship target missing: {to_table}.{to_col}")

    if errors:
        raise RuntimeError("Invalid semantic model structure:\n" + "\n".join(errors))


def validate_visual_bindings(project: Path) -> None:
    """Fail generation if PBIR visuals point at missing model fields."""
    table_dir = project / "Order2Cash.SemanticModel" / "definition" / "tables"
    model: dict[str, dict[str, set[str]]] = {}
    for table_file in table_dir.glob("*.tmdl"):
        text = table_file.read_text(encoding="utf-8")
        table_match = re.search(r"^table\s+(.+)$", text, flags=re.MULTILINE)
        if not table_match:
            continue
        table_name = table_match.group(1).strip().strip("'")
        columns = {match.group(1).strip().strip("'") for match in re.finditer(r"^\s+column\s+(.+)$", text, flags=re.MULTILINE)}
        measures = {match.group(1).strip() for match in re.finditer(r"^\s+measure\s+'([^']+)'", text, flags=re.MULTILINE)}
        model[table_name] = {"columns": columns, "measures": measures}

    allowed_visual_types = {"card", "clusteredBarChart", "clusteredColumnChart", "tableEx"}
    invalid_visual_types = {"cardVisual", "barChart", "columnChart"}
    errors: list[str] = []
    visual_root = project / "Order2Cash.Report" / "definition" / "pages"
    for visual_file in visual_root.glob("**/visual.json"):
        visual = json.loads(visual_file.read_text(encoding="utf-8"))
        visual_type = visual.get("visual", {}).get("visualType")
        if visual_type in invalid_visual_types or visual_type not in allowed_visual_types:
            errors.append(f"{visual_file}: unsupported visualType {visual_type!r}")
        query_state = visual.get("visual", {}).get("query", {}).get("queryState", {})
        if not query_state:
            errors.append(f"{visual_file}: missing queryState")
        for role in query_state.values():
            for projection in role.get("projections", []):
                field = projection.get("field", {})
                measure = field.get("Measure")
                column = field.get("Column")
                if measure:
                    table = measure["Expression"]["SourceRef"]["Entity"]
                    name = measure["Property"]
                    if table not in model or name not in model[table]["measures"]:
                        errors.append(f"{visual_file}: missing measure {table}.{name}")
                if column:
                    table = column["Expression"]["SourceRef"]["Entity"]
                    name = column["Property"]
                    if table not in model or name not in model[table]["columns"]:
                        errors.append(f"{visual_file}: missing column {table}.{name}")

    if errors:
        raise RuntimeError("Invalid generated visual bindings:\n" + "\n".join(errors))


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", type=Path, default=OUT_DIR)
    parser.add_argument(
        "--source-mode",
        choices=["embedded", "native-excel"],
        default="embedded",
        help="embedded writes DATATABLE demo tables for PBIP smoke tests; native-excel writes Power Query Excel.Workbook/File.Contents partitions.",
    )
    args = parser.parse_args()
    global SOURCE_MODE
    SOURCE_MODE = args.source_mode

    if args.out.exists():
        shutil.rmtree(args.out)

    project = args.out
    report = project / "Order2Cash.Report"
    semantic = project / "Order2Cash.SemanticModel"
    report_definition = report / "definition"
    definition = semantic / "definition"
    tables_dir = definition / "tables"

    write(project / ".gitignore", "**/.pbi/localSettings.json\n**/.pbi/cache.abf\n")
    write(
        project / "Order2Cash.pbip",
        json.dumps(
            {
                "version": "1.0",
                "artifacts": [{"report": {"path": "Order2Cash.Report"}}],
                "settings": {"enableAutoRecovery": True},
            },
            indent=2,
        ),
    )
    write(
        report / "definition.pbir",
        json.dumps(
            {
                "$schema": PBIR_SCHEMA,
                "version": "4.0",
                "datasetReference": {"byPath": {"path": "../Order2Cash.SemanticModel"}},
            },
            indent=2,
        ),
    )
    write_report_pages(report_definition)
    write(project / "docs" / "ORDER2CASH_REPORT_SPEC.md", (TEST_PACK / "report" / "ORDER2CASH_REPORT_SPEC.md").read_text(encoding="utf-8"))

    write(
        semantic / "definition.pbism",
        json.dumps(
            {
                "$schema": PBISM_SCHEMA,
                "version": "4.0",
                "settings": {},
            },
            indent=2,
        ),
    )
    write(definition / "database.tmdl", database_tmdl())
    write(definition / "model.tmdl", model_tmdl())
    write(definition / "cultures" / "en-US.tmdl", "cultureInfo en-US\n")
    for table in TABLES:
        write(tables_dir / f"{table}.tmdl", table_tmdl(table))
    write(definition / "relationships.tmdl", relationships_tmdl())

    shutil.copytree(TEST_PACK / "docs", project / "docs", dirs_exist_ok=True)
    shutil.copytree(TEST_PACK / "data", project / "data")
    shutil.copytree(TEST_PACK / "powerquery", project / "powerquery")
    shutil.copy(TEST_PACK / "README.md", project / "README.md")
    with (project / "README.md").open("a", encoding="utf-8") as handle:
        handle.write(
            f"""

## Generated PBIP/PBIR Content

- `Order2Cash.pbip`: Power BI Project entry point.
- `Order2Cash.Report/definition.pbir`: report pointer with relative semantic model reference.
- `Order2Cash.Report/definition/pages`: generated PBIR page folders.
- `Order2Cash.Report/definition/pages/*/visuals/*/visual.json`: generated visual metadata and query placeholders.
- `Order2Cash.SemanticModel/definition`: generated TMDL model, table, measure, relationship, and expression files.

Generated pages: {len(PAGE_SPECS)}
Generated visuals: {sum(len(page['visuals']) for page in PAGE_SPECS)}

Note: Power BI Desktop may normalize or enrich PBIR JSON after opening/saving because PBIR is still a preview/developer format. This scaffold is designed so pages and visuals are always emitted alongside the semantic model.
"""
        )

    validate_tmdl_datatable_constants(project)
    validate_semantic_model_structure(project)
    validate_visual_bindings(project)

    print(f"Wrote PBIP project scaffold to {project}")
    print(f"Open candidate: {project / 'Order2Cash.pbip'}")
    print(f"Report definition: {report / 'definition.pbir'}")
    print(f"Generated PBIR pages: {len(PAGE_SPECS)}")
    print(f"Generated PBIR visuals: {sum(len(page['visuals']) for page in PAGE_SPECS)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
