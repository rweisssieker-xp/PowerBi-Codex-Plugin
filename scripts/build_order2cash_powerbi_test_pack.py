#!/usr/bin/env python3
"""Build a Power BI-ready Order2Cash test pack from the industrial demo workbook."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT / "outputs" / "powerbi-industrial-demo" / "powerbi_industrial_demo_data.xlsx"
OUT_DIR = ROOT / "outputs" / "powerbi-order2cash-test"


def rows(sheet) -> list[dict[str, object]]:
    values = list(sheet.iter_rows(values_only=True))
    headers = [str(h) for h in values[0]]
    return [dict(zip(headers, row)) for row in values[1:] if any(v is not None for v in row)]


def as_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()


def pct(value: float) -> str:
    return f"{value:.1%}"


def eur(value: float) -> str:
    return f"{value:,.0f}".replace(",", ".") + " EUR"


def write(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def table_query(sheet: str, source_path: Path) -> str:
    escaped = str(source_path).replace("\\", "/")
    return f"""// Power Query M - import {sheet} from the industrial demo workbook.
let
    Source = Excel.Workbook(File.Contents("{escaped}"), null, true),
    Raw = Source{{[Item="{sheet}", Kind="Sheet"]}}[Data],
    PromotedHeaders = Table.PromoteHeaders(Raw, [PromoteAllScalars = true])
in
    PromotedHeaders
"""


def csv_query(table_name: str, csv_path: Path) -> str:
    escaped = str(csv_path).replace("\\", "/")
    return f"""// Power Query M - import {table_name} exception table.
let
    Source = Csv.Document(File.Contents("{escaped}"), [Delimiter = ",", Encoding = 65001, QuoteStyle = QuoteStyle.Csv]),
    PromotedHeaders = Table.PromoteHeaders(Source, [PromoteAllScalars = true])
in
    PromotedHeaders
"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=SOURCE_XLSX)
    parser.add_argument("--out", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.source, read_only=True, data_only=True)
    data = {name: rows(wb[name]) for name in wb.sheetnames}

    customers = {r["CustomerID"]: r for r in data["Dim_Customer"]}
    products = {r["ProductID"]: r for r in data["Dim_Product"]}
    sales_orders = data["Fact_SalesOrders"]
    deliveries = data["Fact_Deliveries"]
    billings = data["Fact_Billing"]
    ar = data["Fact_AR"]
    quotes = data["Fact_Quotes"]
    service = data["Fact_ServiceTickets"]

    order_value = sum(float(r["OrderValue"]) for r in sales_orders)
    open_orders = [r for r in sales_orders if r["Status"] in {"Open", "Partially Delivered"}]
    atp_issues = [r for r in sales_orders if r["ATPStatus"] != "Confirmed"]
    otif_bad = [r for r in deliveries if r["OTIFStatus"] != "OTIF"]
    billing_blocks = [r for r in billings if bool(r["BillingBlockFlag"])]
    overdue_ar = [r for r in ar if float(r["OpenAmount"] or 0) > 0]
    disputes = [r for r in ar if bool(r["DisputeFlag"])]
    quote_value = sum(float(r["QuoteValue"]) for r in quotes)
    discount_leakage = sum(float(r["QuoteValue"]) * float(r["DiscountPct"]) / 100 for r in quotes)

    last_order_by_customer: dict[str, date] = {}
    revenue_by_customer: defaultdict[str, float] = defaultdict(float)
    late_by_customer: defaultdict[str, int] = defaultdict(int)
    dispute_by_customer: defaultdict[str, int] = defaultdict(int)
    service_by_customer: defaultdict[str, int] = defaultdict(int)
    for so in sales_orders:
        cid = str(so["CustomerID"])
        od = as_date(so["OrderDate"])
        if od and (cid not in last_order_by_customer or od > last_order_by_customer[cid]):
            last_order_by_customer[cid] = od
        revenue_by_customer[cid] += float(so["OrderValue"])
        if so["ATPStatus"] != "Confirmed":
            late_by_customer[cid] += 1
    for d in deliveries:
        if d["OTIFStatus"] != "OTIF":
            so = next((x for x in sales_orders if x["SalesOrderID"] == d["SalesOrderID"]), None)
            if so:
                late_by_customer[str(so["CustomerID"])] += 1
    for row in ar:
        if row["DisputeFlag"]:
            dispute_by_customer[str(row["CustomerID"])] += 1
    for row in service:
        if row["SLAStatus"] in {"At Risk", "Breached"}:
            service_by_customer[str(row["CustomerID"])] += 1

    as_of = date(2026, 5, 3)
    churn_rows = []
    for cid, customer in customers.items():
        last_order = last_order_by_customer.get(cid)
        days_since = (as_of - last_order).days if last_order else 999
        risk_score = 0
        risk_score += 35 if days_since > 180 else 15 if days_since > 120 else 0
        risk_score += min(25, late_by_customer[cid] * 5)
        risk_score += min(20, dispute_by_customer[cid] * 10)
        risk_score += min(20, service_by_customer[cid] * 7)
        churn_rows.append(
            {
                "CustomerID": cid,
                "CustomerName": customer["CustomerName"],
                "Segment": customer["Segment"],
                "Revenue": revenue_by_customer[cid],
                "LastOrderDate": last_order.isoformat() if last_order else "none",
                "DaysSinceLastOrder": days_since,
                "LateOrShortEvents": late_by_customer[cid],
                "Disputes": dispute_by_customer[cid],
                "ServiceIssues": service_by_customer[cid],
                "ChurnRiskScore": risk_score,
                "RiskBand": "High" if risk_score >= 50 else "Medium" if risk_score >= 25 else "Low",
            }
        )
    churn_rows.sort(key=lambda r: (r["ChurnRiskScore"], r["Revenue"]), reverse=True)
    high_risk = [r for r in churn_rows if r["RiskBand"] == "High"]

    top_products = defaultdict(float)
    for so in sales_orders:
        top_products[str(so["ProductID"])] += float(so["OrderValue"])
    product_rows = sorted(top_products.items(), key=lambda x: x[1], reverse=True)[:10]

    exceptions: list[dict[str, object]] = []
    for row in churn_rows:
        if row["RiskBand"] in {"High", "Medium"}:
            exceptions.append(
                {
                    "ExceptionID": f"EX-CHURN-{row['CustomerID']}",
                    "ProblemType": "Customer Churn Risk",
                    "Severity": row["RiskBand"],
                    "CustomerID": row["CustomerID"],
                    "ProductID": "",
                    "RelatedDocument": row["CustomerID"],
                    "EventDate": row["LastOrderDate"] if row["LastOrderDate"] != "none" else as_of.isoformat(),
                    "AmountExposure": round(float(row["Revenue"]), 2),
                    "Reason": f"Risk score {row['ChurnRiskScore']}; {row['DaysSinceLastOrder']} days since last order; late/short={row['LateOrShortEvents']}; disputes={row['Disputes']}",
                    "RecommendedAction": "Account owner retention call, service recovery, delivery review, commercial offer validation",
                    "OwnerRole": "Sales / Customer Success",
                }
            )
    for so in atp_issues:
        exceptions.append(
            {
                "ExceptionID": f"EX-ATP-{so['SalesOrderID']}",
                "ProblemType": "ATP / Availability Issue",
                "Severity": "High" if so["ATPStatus"] == "Shortage" else "Medium",
                "CustomerID": so["CustomerID"],
                "ProductID": so["ProductID"],
                "RelatedDocument": so["SalesOrderID"],
                "EventDate": so["OrderDate"],
                "AmountExposure": round(float(so["OrderValue"]), 2),
                "Reason": f"ATP status {so['ATPStatus']} for requested date {so['RequestedDate']}",
                "RecommendedAction": "Check constrained supply, substitute product, expedite production or reset confirmed date",
                "OwnerRole": "Supply Chain / Order Management",
            }
        )
    for d in otif_bad:
        so = next((x for x in sales_orders if x["SalesOrderID"] == d["SalesOrderID"]), None)
        exceptions.append(
            {
                "ExceptionID": f"EX-OTIF-{d['DeliveryID']}",
                "ProblemType": "Late Or Short Delivery",
                "Severity": "High",
                "CustomerID": d["CustomerID"],
                "ProductID": d["ProductID"],
                "RelatedDocument": d["DeliveryID"],
                "EventDate": d["ShipDate"],
                "AmountExposure": round(float(so["OrderValue"]) if so else 0, 2),
                "Reason": f"OTIF status {d['OTIFStatus']} via carrier {d['Carrier']}",
                "RecommendedAction": "Review warehouse pick/pack, carrier performance, shortage reason and customer communication",
                "OwnerRole": "Warehouse / Logistics",
            }
        )
    for b in billing_blocks:
        exceptions.append(
            {
                "ExceptionID": f"EX-BILL-{b['InvoiceID']}",
                "ProblemType": "Billing Block",
                "Severity": "High",
                "CustomerID": b["CustomerID"],
                "ProductID": "",
                "RelatedDocument": b["InvoiceID"],
                "EventDate": b["InvoiceDate"],
                "AmountExposure": round(float(b["InvoiceAmount"]), 2),
                "Reason": "Invoice is blocked and may delay revenue/cash",
                "RecommendedAction": "Resolve pricing, tax, delivery or master-data block and rebill if required",
                "OwnerRole": "Billing / Finance",
            }
        )
    for row in overdue_ar:
        exceptions.append(
            {
                "ExceptionID": f"EX-AR-{row['InvoiceID']}",
                "ProblemType": "Open AR / Cash Risk",
                "Severity": "High" if bool(row["DisputeFlag"]) else "Medium",
                "CustomerID": row["CustomerID"],
                "ProductID": "",
                "RelatedDocument": row["InvoiceID"],
                "EventDate": row["DueDate"],
                "AmountExposure": round(float(row["OpenAmount"] or 0), 2),
                "Reason": f"Open amount with dunning level {row['DunningLevel']}; dispute={row['DisputeFlag']}",
                "RecommendedAction": "Collections follow-up, dispute resolution, credit hold review, root-cause coding",
                "OwnerRole": "Accounts Receivable",
            }
        )
    exceptions.sort(key=lambda r: float(r["AmountExposure"]), reverse=True)

    kpis = {
        "Sales order count": len(sales_orders),
        "Order value": eur(order_value),
        "Open backlog value": eur(sum(float(r["OrderValue"]) for r in open_orders)),
        "ATP issue rate": pct(len(atp_issues) / len(sales_orders)),
        "Late or short delivery rate": pct(len(otif_bad) / len(deliveries)),
        "Billing block rate": pct(len(billing_blocks) / len(billings)),
        "Open AR amount": eur(sum(float(r["OpenAmount"] or 0) for r in overdue_ar)),
        "Dispute rate": pct(len(disputes) / len(ar)),
        "Estimated discount leakage": eur(discount_leakage),
        "High churn-risk customers": len(high_risk),
        "O2C exception count": len(exceptions),
        "O2C exception exposure": eur(sum(float(r["AmountExposure"]) for r in exceptions)),
    }

    powerquery_dir = args.out / "powerquery"
    dax_dir = args.out / "dax"
    docs_dir = args.out / "docs"
    report_dir = args.out / "report"
    model_dir = args.out / "model"
    for sheet in [
        "Dim_Calendar",
        "Dim_Customer",
        "Dim_Product",
        "Dim_Plant",
        "Dim_Warehouse",
        "Fact_Leads",
        "Fact_Opportunities",
        "Fact_Quotes",
        "Fact_SalesOrders",
        "Fact_Deliveries",
        "Fact_Billing",
        "Fact_AR",
        "Fact_ServiceTickets",
        "Fact_Shipments",
        "Fact_GL_Postings",
    ]:
        write(powerquery_dir / f"{sheet}.pq", table_query(sheet, args.source))

    data_dir = args.out / "data"
    exceptions_csv = data_dir / "Fact_Order2CashExceptions.csv"
    data_dir.mkdir(parents=True, exist_ok=True)
    with exceptions_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(exceptions[0].keys()))
        writer.writeheader()
        writer.writerows(exceptions)
    write(powerquery_dir / "Fact_Order2CashExceptions.pq", csv_query("Fact_Order2CashExceptions", exceptions_csv))

    measures = """-- Order2Cash Power BI measures for the industrial demo workbook.

Order Count =
COUNTROWS ( Fact_SalesOrders )

Order Value =
SUM ( Fact_SalesOrders[OrderValue] )

Open Backlog Value =
CALCULATE (
    [Order Value],
    Fact_SalesOrders[Status] IN { "Open", "Partially Delivered" }
)

ATP Issue Count =
CALCULATE (
    [Order Count],
    Fact_SalesOrders[ATPStatus] <> "Confirmed"
)

ATP Issue Rate =
DIVIDE ( [ATP Issue Count], [Order Count] )

Delivered Quantity =
SUM ( Fact_Deliveries[DeliveredQty] )

Late Or Short Deliveries =
CALCULATE (
    COUNTROWS ( Fact_Deliveries ),
    Fact_Deliveries[OTIFStatus] <> "OTIF"
)

Late Or Short Delivery Rate =
DIVIDE ( [Late Or Short Deliveries], COUNTROWS ( Fact_Deliveries ) )

Invoice Amount =
SUM ( Fact_Billing[InvoiceAmount] )

Billing Block Count =
CALCULATE (
    COUNTROWS ( Fact_Billing ),
    Fact_Billing[BillingBlockFlag] = TRUE ()
)

Billing Block Rate =
DIVIDE ( [Billing Block Count], COUNTROWS ( Fact_Billing ) )

Open AR Amount =
SUM ( Fact_AR[OpenAmount] )

Dispute Count =
CALCULATE (
    COUNTROWS ( Fact_AR ),
    Fact_AR[DisputeFlag] = TRUE ()
)

O2C Exception Count =
COUNTROWS ( Fact_Order2CashExceptions )

O2C Exception Exposure =
SUM ( Fact_Order2CashExceptions[AmountExposure] )

High Severity Exceptions =
CALCULATE (
    [O2C Exception Count],
    Fact_Order2CashExceptions[Severity] = "High"
)

Dunning Customers =
CALCULATE (
    DISTINCTCOUNT ( Fact_AR[CustomerID] ),
    Fact_AR[DunningLevel] > 0
)

Quote Value =
SUM ( Fact_Quotes[QuoteValue] )

Discount Leakage =
SUMX ( Fact_Quotes, Fact_Quotes[QuoteValue] * Fact_Quotes[DiscountPct] / 100 )

Churn Risk Customers =
CALCULATE (
    DISTINCTCOUNT ( Dim_Customer[CustomerID] ),
    Dim_Customer[ActiveFlag] = FALSE ()
)

Customer Churn Risk Score =
VAR DaysSinceLastOrder =
    DATEDIFF (
        CALCULATE ( MAX ( Fact_SalesOrders[OrderDate] ) ),
        TODAY (),
        DAY
    )
VAR LateEvents =
    CALCULATE ( [ATP Issue Count] + [Late Or Short Deliveries] )
VAR Disputes = [Dispute Count]
RETURN
    MIN (
        100,
        IF ( DaysSinceLastOrder > 180, 35, IF ( DaysSinceLastOrder > 120, 15, 0 ) )
            + MIN ( 25, LateEvents * 5 )
            + MIN ( 20, Disputes * 10 )
    )
"""
    write(dax_dir / "order2cash_measures.dax", measures)

    relationships = """# Order2Cash Semantic Model Relationships

## Dimensions

- `Dim_Calendar[Date]` to role-playing dates in orders, quotes, deliveries, billings, AR, service tickets, and shipments.
- `Dim_Customer[CustomerID]` -> facts with `CustomerID`.
- `Dim_Product[ProductID]` -> sales orders, deliveries, service tickets.
- `Dim_Plant[PlantID]` -> sales orders.
- `Dim_Warehouse[WarehouseID]` -> deliveries.

## Fact Flow

- `Fact_Leads[LeadID]` -> `Fact_Opportunities[LeadID]`
- `Fact_Opportunities[OpportunityID]` -> `Fact_Quotes[OpportunityID]`
- `Fact_Quotes[QuoteID]` -> `Fact_SalesOrders[QuoteID]`
- `Fact_SalesOrders[SalesOrderID]` -> `Fact_Deliveries[SalesOrderID]`
- `Fact_Deliveries[DeliveryID]` -> `Fact_Billing[DeliveryID]`
- `Fact_Billing[InvoiceID]` -> `Fact_AR[InvoiceID]`
- `Fact_Deliveries[DeliveryID]` -> `Fact_Shipments[DeliveryID]`
- `Fact_Order2CashExceptions[CustomerID]` -> `Dim_Customer[CustomerID]`
- `Fact_Order2CashExceptions[ProductID]` -> `Dim_Product[ProductID]` where populated

## Notes

- Keep the main customer/product/date dimensions as one-to-many relationships.
- Keep fact-to-fact relationships inactive or use bridge/detail pages if Power BI creates ambiguity.
- Prefer a process-event table later if true process mining is needed.
"""
    write(model_dir / "relationships.md", relationships)

    report_spec = f"""# Order2Cash Power BI Report Spec

## Source

`{args.source}`

## Business Goal

Show all major Order2Cash problems in one Power BI cockpit: customer churn risk, ATP shortages, backlog, late/short delivery, billing blocks, open receivables, disputes, discount leakage, and service/SLA risks.

## Pages

1. **O2C Executive Overview**: KPI cards, process health score, top 10 exceptions, financial exposure.
2. **Customer Churn and Retention Risk**: churn-risk score, no-order aging, disputes, delivery failures, service issues.
3. **Order Intake and Backlog**: order value, open backlog, ATP status, requested vs confirmed dates.
4. **Delivery and OTIF**: late/short delivery rate, carrier performance, warehouse and plant root causes.
5. **Billing and Cash**: billing blocks, invoice amount, open AR, dunning level, disputes.
6. **Quote and Margin Leakage**: discount leakage, approval cycle days, legal loops, quote conversion.
7. **Exception Action Backlog**: customer, owner, problem type, estimated exposure, recommended action.

## Demo KPIs From Current Workbook

| KPI | Value |
| --- | --- |
{chr(10).join(f"| {k} | {v} |" for k, v in kpis.items())}

## Top Churn-Risk Customers

| Customer | Segment | Revenue | Days Since Last Order | Late/Short Events | Disputes | Risk Score | Risk Band |
| --- | --- | ---: | ---: | ---: | ---: | ---: | --- |
{chr(10).join(f"| {r['CustomerName']} | {r['Segment']} | {eur(r['Revenue'])} | {r['DaysSinceLastOrder']} | {r['LateOrShortEvents']} | {r['Disputes']} | {r['ChurnRiskScore']} | {r['RiskBand']} |" for r in churn_rows[:15])}

## Top Products By Order Value

| Product | Product Family | Order Value |
| --- | --- | ---: |
{chr(10).join(f"| {products[pid]['ProductName']} | {products[pid]['ProductFamily']} | {eur(value)} |" for pid, value in product_rows)}

## Top O2C Exceptions By Exposure

| Problem | Severity | Customer | Related Document | Exposure | Recommended Action |
| --- | --- | --- | --- | ---: | --- |
{chr(10).join(f"| {r['ProblemType']} | {r['Severity']} | {customers.get(r['CustomerID'], {}).get('CustomerName', r['CustomerID'])} | {r['RelatedDocument']} | {eur(float(r['AmountExposure']))} | {r['RecommendedAction']} |" for r in exceptions[:15])}
"""
    write(report_dir / "ORDER2CASH_REPORT_SPEC.md", report_spec)

    readme = f"""# Power BI Order2Cash Test Pack

This pack uses the generated industrial Excel workbook as a Power BI source and provides a complete Order2Cash implementation blueprint.

## Contents

- `powerquery/*.pq`: one Power Query M import per required sheet.
- `data/Fact_Order2CashExceptions.csv`: consolidated action backlog for all detected O2C problems.
- `dax/order2cash_measures.dax`: KPI measures for O2C problems and customer churn risk.
- `model/relationships.md`: semantic model relationship plan.
- `report/ORDER2CASH_REPORT_SPEC.md`: report pages, KPI values from the demo workbook, and top-risk customers.
- `docs/order2cash.en-US.md` and `docs/order2cash.de-DE.md`: bilingual business documentation.

## Source Workbook

`{args.source}`

## How To Use In Power BI Desktop

1. Open Power BI Desktop.
2. Import the Excel workbook.
3. Load the sheets listed in `powerquery/`.
4. Create relationships from `model/relationships.md`.
5. Add measures from `dax/order2cash_measures.dax`.
6. Build pages from `report/ORDER2CASH_REPORT_SPEC.md`.

## Test Outcome

The current demo data produces {len(high_risk)} high-risk churn customers, {len(exceptions)} O2C exception rows, {pct(len(atp_issues) / len(sales_orders))} ATP issue rate, {pct(len(otif_bad) / len(deliveries))} late/short delivery rate, and {eur(sum(float(r["OpenAmount"] or 0) for r in overdue_ar))} open AR amount.
"""
    write(args.out / "README.md", readme)

    en = """# Order2Cash Problem Cockpit Documentation

The report identifies O2C problems across customer demand, order confirmation, delivery execution, billing, cash collection, and customer churn risk.

Primary questions:

- Which customers are at churn risk because of no recent orders, delivery failures, disputes, or service issues?
- Which orders are blocked by ATP shortages or delayed confirmations?
- Which deliveries are late or short and which carriers, plants, or warehouses drive the issue?
- Which invoices are blocked, disputed, overdue, or driving dunning actions?
- Which quotes create discount leakage before the order is booked?

Recommended action model: every exception should have owner, due date, estimated financial exposure, recommended action, and retest condition.
"""
    de = """# Order2Cash Problem-Cockpit Dokumentation

Der Report zeigt O2C-Probleme ueber Kundennachfrage, Auftragsbestaetigung, Lieferung, Faktura, Zahlungseingang und Customer-Churn-Risiko.

Zentrale Fragen:

- Welche Kunden haben Churn-Risiko durch lange Bestellpause, Lieferprobleme, Disputes oder Serviceprobleme?
- Welche Auftraege sind durch ATP-Engpaesse oder spaete Bestaetigungen gefaehrdet?
- Welche Lieferungen sind verspaetet oder unvollstaendig und welche Carrier, Werke oder Lager treiben das Problem?
- Welche Rechnungen sind blockiert, disputiert, ueberfaellig oder im Mahnprozess?
- Welche Angebote erzeugen Rabatt- und Margenleckage vor Auftragseingang?

Empfohlenes Aktionsmodell: jede Ausnahme braucht Owner, Faelligkeit, finanzielles Exposure, empfohlene Massnahme und Retest-Bedingung.
"""
    write(docs_dir / "order2cash.en-US.md", en)
    write(docs_dir / "order2cash.de-DE.md", de)

    print(f"Wrote Order2Cash Power BI test pack to {args.out}")
    print(f"High churn-risk customers: {len(high_risk)}")
    print(f"O2C exception rows: {len(exceptions)}")
    print(f"ATP issue rate: {pct(len(atp_issues) / len(sales_orders))}")
    print(f"Late/short delivery rate: {pct(len(otif_bad) / len(deliveries))}")
    print(f"Open AR: {eur(sum(float(r['OpenAmount'] or 0) for r in overdue_ar))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
